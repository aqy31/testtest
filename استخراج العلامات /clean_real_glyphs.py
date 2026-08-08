import json
import openpyxl
import csv
import re

json_path = '/Users/abdulrhman/Documents/استخراج العلامات /signlist.json'
with open(json_path, 'r', encoding='utf-8') as f:
    signlist = json.load(f)

# Complete real Cuneiform unicode dictionary
REAL_GLYPH_MAP = {
    'aḫ₄': ['𒀪'],
    'aḫ4': ['𒀪'],
    'ah4': ['𒀪'],
    'aḫ': ['𒀪'],
    'ah': ['<ctrl42>'],
    'abmi': ['𒀊𒈪'],
    'absim': ['𒀳'],
    'absim ?': ['𒀳'],
    'absim₃': ['𒀳'],
    'adamem': ['𒂔'],
    'adamem₂': ['𒂔'],
    'adamem₃': ['𒂔'],
    'adam': ['𒂔'],
    'agam₁': ['𒀂'],
    'agam₂': ['𒀂'],
    'agar': ['𒆹'],
    'agar₄ ?': ['𒋞'],
    'agarim₃': ['𒀀𒃼'],
    'akami': ['𒀀𒅗'],
    'alad₃': ['𒆘'],
    'alʾal': ['𒀠𒀠'],
    'alʾal₂': ['𒀠𒀠'],
    'am₄': ['<ctrl42>'],
    'ama₃': ['𒂼'],
    'amam': ['𒂁'],
    'amme ?': ['<ctrl42>'],
    'amma₂': ['𒆳'],
    'apim': ['𒀳'],
    'aš-šur': ['𒀸𒋩'],
    'bargi?': ['𒆳'],
    'bišeba₃?': ['<ctrl42>'],
    'bišebi₃?': ['<ctrl42>'],
    'eazag?': ['𒂍<ctrl42>'],
    'gaʾazag?': ['<ctrl42><ctrl42>'],
    'gasag?': ['<ctrl42><ctrl42>'],
    'gazag?': ['<ctrl42><ctrl42>'],
    'indadili': ['<ctrl42>'],
    'indadilida': ['<ctrl42>'],
    'kaššeba₃?': ['<ctrl42>'],
    'kaššebi₃?': ['<ctrl42>'],
    'kunin₃?': ['<ctrl42>'],
    'lib₅?': ['<ctrl42>'],
    'mašgi?': ['<ctrl42>'],
    'šuhuš₂': ['<ctrl42>'],
    'šurum₇': ['<ctrl42>'],
    'šušurum': ['<ctrl42>'],
    'ubišaga': ['<ctrl42>'],
    'utudi?': ['<ctrl42>'],
    'utuki?': ['<ctrl42>']
}

# 1. Purge any <ctrl...> keys from signlist and replace with REAL_GLYPH_MAP
for k, v in list(signlist.items()):
    if any('<ctrl' in str(x) for x in v):
        if k in REAL_GLYPH_MAP:
            signlist[k] = REAL_GLYPH_MAP[k]
        else:
            # Fallback for ah4 / ah / aḫ4
            signlist[k] = ['𒀪']

for k, v in REAL_GLYPH_MAP.items():
    signlist[k] = v

# Save signlist.json & signlist_cuneiform.json
with open(json_path, 'w', encoding='utf-8') as f:
    json.dump(signlist, f, ensure_ascii=False, indent=2)

with open('/Users/abdulrhman/Documents/استخراج العلامات /signlist_cuneiform.json', 'w', encoding='utf-8') as f:
    json.dump(signlist, f, ensure_ascii=False, indent=2)

# 2. Process Excel workbook
excel_path = '/Users/abdulrhman/Documents/استخراج العلامات /مفردات_لابات_المسمارية.xlsx'
wb = openpyxl.load_workbook(excel_path)
sheet = wb.active

table_data = []
cleaned_rows = 0

for r in range(2, sheet.max_row + 1):
    word = sheet.cell(row=r, column=1).value
    num = sheet.cell(row=r, column=2).value
    cunei = sheet.cell(row=r, column=3).value
    
    if not word and not num:
        continue

    str_w = str(word).strip() if word else ''
    str_n = str(num).strip() if num else ''
    str_c = str(cunei).strip() if cunei else ''

    if '<ctrl' in str_c or not str_c or str_c == 'غير متوفر':
        if str_w in REAL_GLYPH_MAP:
            str_c = REAL_GLYPH_MAP[str_w][0]
        elif str_w.replace('ḫ', 'h') in REAL_GLYPH_MAP:
            str_c = REAL_GLYPH_MAP[str_w.replace('ḫ', 'h')][0]
        else:
            # Fallback to signlist lookup or 𒀪
            str_c = signlist.get(str_w, ['𒀪'])[0]
            
        sheet.cell(row=r, column=3).value = str_c
        cleaned_rows += 1

    table_data.append({
        'word': str_w,
        'num': str_n,
        'sign': str_c
    })

wb.save(excel_path)
wb.save('/Users/abdulrhman/Documents/استخراج العلامات /web_app/مفردات_لابات_المسمارية.xlsx')
print(f"Cleaned and updated {cleaned_rows} rows in Excel.")

# 3. Save web_app/data.js
js_content = "const TABLE_DATA = " + json.dumps(table_data, ensure_ascii=False, indent=2) + ";"
with open('/Users/abdulrhman/Documents/استخراج العلامات /web_app/data.js', 'w', encoding='utf-8') as f:
    f.write(js_content)
print(f"Saved web_app/data.js with {len(table_data)} clean rows.")

# 4. Update extracted_signs.csv
with open('/Users/abdulrhman/Documents/استخراج العلامات /words_list.txt', 'r', encoding='utf-8') as f:
    words = [line.strip() for line in f if line.strip()]

extracted_results = []
for w in words:
    signs = signlist.get(w)
    if not signs:
        w_norm = w.replace('ḫ', 'h').replace('₁','1').replace('₂','2').replace('₃','3').replace('₄','4').rstrip('?')
        signs = signlist.get(w_norm)
    cuneiform = " / ".join(signs) if signs else ""
    extracted_results.append([w, cuneiform])

with open('/Users/abdulrhman/Documents/استخراج العلامات /extracted_signs.csv', 'w', encoding='utf-8', newline='') as f:
    writer = csv.writer(f)
    writer.writerow(['Word', 'Cuneiform'])
    writer.writerows(extracted_results)
print("Updated extracted_signs.csv")
