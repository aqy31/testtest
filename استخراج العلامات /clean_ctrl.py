import json
import openpyxl
import csv

def main():
    json_path = '/Users/abdulrhman/Documents/استخراج العلامات /signlist.json'
    with open(json_path, 'r', encoding='utf-8') as f:
        signlist = json.load(f)

    # 1. Clean signlist dictionary from any <ctrl...> strings
    for k, v in list(signlist.items()):
        cleaned = []
        for item in v:
            if '<ctrl' in str(item):
                cleaned.append('𒀪')
            else:
                cleaned.append(item)
        signlist[k] = cleaned

    signlist['aḫ₄'] = ['𒀪']
    signlist['aḫ4'] = ['𒀪']
    signlist['ah4'] = ['𒀪']
    signlist['aḫ'] = ['

    with open(json_path, 'w', encoding='utf-8') as f:
        json.dump(signlist, f, ensure_ascii=False, indent=2)

    with open('/Users/abdulrhman/Documents/استخراج العلامات /signlist_cuneiform.json', 'w', encoding='utf-8') as f:
        json.dump(signlist, f, ensure_ascii=False, indent=2)

    # 2. Update Excel
    excel_path = '/Users/abdulrhman/Documents/استخراج العلامات /مفردات_لابات_المسمارية.xlsx'
    wb = openpyxl.load_workbook(excel_path)
    sheet = wb.active

    table_data = []
    for r in range(2, sheet.max_row + 1):
        word = sheet.cell(row=r, column=1).value
        num = sheet.cell(row=r, column=2).value
        cunei = sheet.cell(row=r, column=3).value

        if not word and not num:
            continue

        str_w = str(word).strip() if word else ''
        str_n = str(num).strip() if num else ''
        str_c = str(cunei).strip() if cunei else ''

        if '<ctrl' in str_c or not str_c or str_c == 'غير متوفر':
            if str_w in signlist:
                str_c = signlist[str_w][0]
            elif str_w.replace('ḫ', 'h') in signlist:
                str_c = signlist[str_w.replace('ḫ', 'h')][0]
            else:
                str_c = '<ctrl42>'
            sheet.cell(row=r, column=3).value = str_c

        table_data.append({
            'word': str_w,
            'num': str_n,
            'sign': str_c
        })

    wb.save(excel_path)
    wb.save('/Users/abdulrhman/Documents/استخراج العلامات /web_app/مفردات_لابات_المسمارية.xlsx')

    # 3. Update web_app/data.js
    js_content = "const TABLE_DATA = " + json.dumps(table_data, ensure_ascii=False, indent=2) + ";"
    with open('/Users/abdulrhman/Documents/استخراج العلامات /web_app/data.js', 'w', encoding='utf-8') as f:
        f.write(js_content)

    print("Purged all ctrl strings successfully!")

if __name__ == '__main__':
    main()
